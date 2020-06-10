import openpyxl
import os
import numpy

os.chdir('I:\\WeatherDataProcessing')

# 1 - Indique a planilha de dados
wb = openpyxl.load_workbook('example_01.xlsx')
plan = wb['ES_A617_ALEGRE_fornecimento_1']

# cria uma lista vazia para armazenar os dados de cada variável em cada mês
listaDados = []

# Para contrar quantas dias de dados existem
contaLinhas = 11
celula = "082018"
while celula != None:
    contaLinhas = contaLinhas + 1
    celula = plan.cell(row=contaLinhas, column=1).value

print('O total de linhas é: ', contaLinhas)

# Para contar quantas variávei existem
contaColunas = 0
coluna = 2
celula = "082018"
while celula != None:
    celula = plan.cell(row=10, column=coluna).value
    coluna = coluna + 1
    contaColunas = contaColunas + 1

print('O total de colunas é: ', contaColunas)

listaVariaveis = []
dic = {}
soma = numpy.sum
media = numpy.mean
minima = numpy.amin
maxima = numpy.max

for i in range(2, contaColunas, 24):
    celula = plan.cell(row=10, column=i).value
    listaVariaveis.append(str(celula))
    if celula == 'TEMPERATURA DO AR (°C)' or celula == 'UMIDADE RELATIVA DO AR (%)' \
            or celula == 'TEMPERATURA DO PONTO DE ORVALHO (°C)':
        j = media
    elif celula == 'TEMPERATURA MÁXIMA   ' or celula == 'TEMPERATURA MÁXIMA DO PONTO DE ORVALHO (°C)' \
            or celula == 'UMIDADE RELATIVA MAXIMA DO AR (%)':
        j = maxima
    elif celula == 'TEMPERATURA MÍNIMA   ' or celula == 'TEMPERATURA MÍNIMA DO PONTO DE ORVALHO (°C)' \
            or celula == 'UMIDADE RELATIVA MINIMA DO AR (%)':
        j = minima
    else:
        j = soma
    dic.update({str(celula): j})

for u, v, (x, z) in zip(range(2, contaColunas+1, 24), range(2, 11, 1), dic.items()):

    listaDados = []
    linha = 2
    colIni = u
    colEnd = colIni + 24

    for i in range(12, 409, 1):
        print('A coluna (u) inicial de dados que está sendo trabalhada é: ', u)
        print('Os dados estão sendo capiturados da linha: ', i)

        # 3 - Indique a planilha de dados
        wb = openpyxl.load_workbook('example_01.xlsx')
        plan = wb['ES_A617_ALEGRE_fornecimento_1']

        if i == 12:
            for j in range(colIni, colEnd, 1):
                temperaturaDia = plan.cell(row=i, column=j).value
                if temperaturaDia != 'NULL':
                    listaDados.append(temperaturaDia)

        elif i != 12:
            # Se i for diferente de 12 ele vai guardar duas datas
            # A data da linha de coleta de dados atual
            try:
                data_01 = plan.cell(row=i - 1, column=1).value
                data_01_n = str(data_01.strftime('%m/%Y'))
                print('Estamos na linha', i)
                print('O valor da data anterior é: ', data_01_n)

                data = plan.cell(row=i, column=1).value
                data_n = str(data.strftime('%m/%Y'))

                print('O valor da data atual é: ', data_n)

            except:
                data_01 = plan.cell(row=i - 1, column=1).value
                data_01_n = str(data_01.strftime('%m/%Y'))

                print(listaDados)
                mes = data_01_n

                calc = z
                media = calc(listaDados)
                pastaDados = openpyxl.load_workbook('danilo.xlsx')
                sheet = pastaDados['Sheet']

                sheet.cell(row=1, column=v).value = x
                sheet.cell(row=linha, column=1).value = mes
                sheet.cell(row=linha, column=v).value = media

                pastaDados.save('danilo.xlsx')

            else:
                if data_n == data_01_n:
                    for j in range(colIni, colEnd, 1):
                        temperaturaDia = plan.cell(row=i, column=j).value
                        if temperaturaDia != 'NULL':
                            listaDados.append(temperaturaDia)

                else:
                    print(listaDados)
                    mes = data_01_n

                    calc = z
                    media = calc(listaDados)
                    pastaDados = openpyxl.load_workbook('danilo.xlsx')
                    sheet = pastaDados['Sheet']

                    sheet.cell(row=1, column=v).value = x
                    sheet.cell(row=linha, column=1).value = mes
                    sheet.cell(row=linha, column=v).value = media

                    linha = linha + 1
                    pastaDados.save('danilo.xlsx')

                    print('O Resultado é : ', numpy.mean(listaDados))
                    print('A data da ultima linha de dados é: ', data_01.strftime('%d/%m/%Y'))
                    listaDados = []
                    print('Zerou a lista na linha de valor igual a :', i)
                    for j in range(colIni, colEnd, 1):
                        temperaturaDia = plan.cell(row=i, column=j).value
                        if temperaturaDia != 'NULL':
                            listaDados.append(temperaturaDia)